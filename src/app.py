import pandas as pd
from datetime import datetime
from dateutil.parser import parse
import re
from unidecode import unidecode
import matplotlib.pyplot as pl
import seaborn as sns

import openpyxl
from openpyxl import load_workbook, Workbook

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage

# Carregar planilha de relação
planilha_relacao = load_workbook(r"C:\Users\henri\Documents\python_projects\RPA\RPA\src\Relacao_Produtos_e_Clientes_2024.xlsx")
planilha = planilha_relacao.active

#Colunas da planilha
coluna_data_venda = "A"
coluna_produto = "B"
coluna_valor_venda = "C"
coluna_regiao = "D"
coluna_equipe_venda = "E"
coluna_cliente = "F"
coluna_metodo_pagamento = "G"
coluna_desconto = "H"

#Lista das colunas
lista_data_venda = []
datas_invalidas = []
lista_produtos = []
lista_valor_venda = []
vendas_invalidas = []
lista_regiao = []
lista_equipe_venda = []
equipes_invalidas = []
lista_cliente = []
lista_metodo_pagamento = []
lista_desconto = []

#PALVRAS CHAVE
debito = ['deb','debito','cartao de debito']
credito = ['cred','credito','cartao de credito']
transferencia_bancaria = ['tran','trans','transferencia','ban','bancaria']
dinheiro = ['din','cash','dinheiro']
cheque = ['cheque','cheq']

def salvar_grafico_e_enviar_email(fig, nome_arquivo, destinatario):
    # Salvar o gráfico como uma imagem
    fig.savefig(nome_arquivo)
    
    # Configurações de e-mail
    remetente = 'henriqueataide.dev@gmail.com'  # Insira o e-mail remetente
    senha = 'hsa$1602'  # Insira a senha do e-mail remetente
    assunto = 'Gráfico'
    corpo = 'Segue em anexo o gráfico solicitado.'

    # Configuração do e-mail
    msg = MIMEMultipart()
    msg['De'] = remetente
    msg['Para'] = destinatario
    msg['Assunto'] = assunto

    # Corpo do e-mail
    msg.attach(MIMEText(corpo, 'plain'))

    # Anexar gráfico
    with open(nome_arquivo, 'rb') as anexo:
        imagem = MIMEImage(anexo.read())
        imagem.add_header('Content-Disposition', 'attachment', filename=nome_arquivo)
        msg.attach(imagem)

# Função para criar gráfico
def grafico(x,y,hue,data,title,xlabel,ylabel):
    pl.figure(figsize=(20, 6))
    sns.barplot(x=x, y=y, hue=hue, data=data, ci=None, orient='x')
    pl.title(title)
    pl.xlabel(xlabel)
    pl.ylabel(ylabel)
    pl.show()

# Valida métodos de pagamentos para separar a frequencia com que são utilizadas
def validar_metodo(celula, metodo_pagamento):
    caracter_limpo = unidecode(celula).lower()    
    for palavra in metodo_pagamento:
        if palavra in caracter_limpo:
            return True
    return False

# Função para validar datas
def validar_data(data_str):
    # Expressão regular para verificar se a entrada se parece com uma data
    regex_data = r'\b\d{1,4}[-/]\d{1,2}[-/]\d{1,4}\b'

    if re.match(regex_data, data_str):
        return True
    else:
        return False

def extrair_valor(num):
    return float (num.replace("$", ""))

# Função para validar se a venda é um número
def validar_numero(num):
    try:
        float(num)
        return True
    except ValueError:
        return False

# Valida para ver se a equipe está dentro do padrão de escrita    
def validar_equipe(equipe):
    equipe = equipe.replace("Equipe ","")  
    try:
        int(equipe)
        return True
    except ValueError:
        return False

# Remove qualquer caracter que não seja um numero da string
def remover_nao_numericos(string):
    return re.sub(r'\D', '', string)
        
# Adicionar a lista as datas que estão validadas
for linha in range(2, planilha.max_row + 1):
    celulas_data_venda = str(planilha[f'{coluna_data_venda}{linha}'].value)
    if validar_data(celulas_data_venda) == True:
        lista_data_venda.append(celulas_data_venda)
    else:
        datas_invalidas.append(celulas_data_venda)
        
# Adicionar os produtos na lista de produtos        
for linha in range(2, planilha.max_row + 1):
    celulas_produto = str(planilha[f'{coluna_produto}{linha}'].value)
    lista_produtos.append(celulas_produto)

# Adicionar valor da venda na lista de valor de venda          
for linha in range(2, planilha.max_row + 1):
    celulas_valor_venda = (planilha[f'{coluna_valor_venda}{linha}'].value)
    celulas_valor_venda = celulas_valor_venda.replace("$", "").replace(",", "")    
    if validar_numero(celulas_valor_venda) == True:
        celulas_valor_venda_num = float(celulas_valor_venda)
        lista_valor_venda.append(celulas_valor_venda_num)
    else:
        vendas_invalidas.append(celulas_valor_venda)

# Adicionar as regiões a lista de região
for linha in range(2, planilha.max_row + 1):
    celulas_regiao = (planilha[f'{coluna_regiao}{linha}'].value)
    lista_regiao.append(celulas_regiao)

# Adicionar as equipes válidas a lista e as invalidas
for linha in range(2, planilha.max_row + 1):
    celulas_equipe_venda = (planilha[f'{coluna_equipe_venda}{linha}'].value)
    if validar_equipe(celulas_equipe_venda) == True:
        lista_equipe_venda.append(celulas_equipe_venda)
    else:
        equipes_invalidas.append(celulas_equipe_venda)

# Adicionar clientes a lista de clientes        
for linha in range(2, planilha.max_row + 1):
    celulas_cliente = (planilha[f'{coluna_cliente}{linha}'].value)
    lista_cliente.append(celulas_cliente)


for linha in range(2, planilha.max_row + 1):
    celulas_metodo_pagamento = (planilha[f'{coluna_metodo_pagamento}{linha}'].value)
    metodo_pagamento_padronizado = unidecode(celulas_metodo_pagamento).lower()
    for debito_chave in debito:
        if debito_chave in metodo_pagamento_padronizado:
            metodo_pagamento_padronizado = "Cartão de Débito"
            break
    for credito_chave in credito:
        if credito_chave in metodo_pagamento_padronizado:
            metodo_pagamento_padronizado = "Cartão de Crédito"
            break
    for transferencia_chave in transferencia_bancaria:
        if transferencia_chave in metodo_pagamento_padronizado:
            metodo_pagamento_padronizado = "Transferência Bancária"
            break
    for dinheiro_chave in dinheiro:
        if dinheiro_chave in metodo_pagamento_padronizado:
            metodo_pagamento_padronizado = "Dinheiro"
            break
    for cheque_chave in cheque:
        if cheque_chave in metodo_pagamento_padronizado:
            metodo_pagamento_padronizado = "Cheque"
            break
        
    lista_metodo_pagamento.append(metodo_pagamento_padronizado)
    
print(lista_metodo_pagamento)

# Adicionar descontos a lista de descontos        
for linha in range(2, planilha.max_row + 1):
    celulas_desconto = str(planilha[f'{coluna_desconto}{linha}'].value)
    celula_num = remover_nao_numericos(celulas_desconto)
    if validar_numero(celula_num) == True:
        lista_desconto.append(celula_num)

# Criando DataFrames a partir das listas que foram atribuídas
data_data_vendas = {
    'Data da Venda': lista_data_venda
}
df_data_vendas = pd.DataFrame(data_data_vendas)

data_produtos = {
    'Produto': lista_produtos
}
df_produtos = pd.DataFrame(data_produtos)

data_valor_venda = {
    'Valor da Venda': lista_valor_venda
}
df_valor_venda = pd.DataFrame(data_valor_venda)

data_regiao = {
    'Região': lista_regiao
}
df_regiao = pd.DataFrame(data_regiao)

data_equipe_venda = {
    'Equipe de Venda': lista_equipe_venda
}
df_equipe_venda = pd.DataFrame(data_equipe_venda)

data_cliente = {
    'Cliente': lista_cliente
}
df_cliente = pd.DataFrame(data_cliente)

# Criar o dicionário com os dados
data_metodo_pagamento = {
    'Método de Pagamento': lista_metodo_pagamento,
}
df_metodo_pagamento = pd.DataFrame(data_metodo_pagamento)

data_desconto = {
    'Desconto': lista_desconto
}
df_desconto = pd.DataFrame(data_desconto)

dfs = [df_produtos, df_valor_venda, df_regiao, df_equipe_venda, df_cliente, df_metodo_pagamento, df_desconto]
df_planilha = pd.concat(dfs, axis=1)  

# Data frame focado em produtos vendidos
df_produtos_vendidos = df_planilha.groupby(['Produto']).value_counts()
df_produtos_vendidos = df_produtos_vendidos.reset_index()
grafico(x='Produto', y='Valor da Venda', hue='Região', data=df_produtos_vendidos,
        title='Distribuição do Valor de Venda por Produto e Região', xlabel='Produto', ylabel='Valor da Venda')

# Data frame focado nas equipes que mais venderam tais produtos por região
df_equipes_de_venda = df_planilha.groupby(['Equipe de Venda']).value_counts()
df_equipes_de_venda = df_equipes_de_venda.reset_index()
grafico(x='Equipe de Venda', y='Valor da Venda', hue='Produto', data=df_equipes_de_venda,
        title='Equipes que mais venderam por Produto', xlabel='Equipes', ylabel='Valor da Venda')

# Data frame focado nos metodos de pagamento mais utilizado pelos clientes
df_metodo_pagamento_clientes = [df_produtos, df_cliente, df_metodo_pagamento]
df_metodo_pagamento_clientes = pd.concat(df_metodo_pagamento_clientes, axis=1)

df_metodo_pagamento_clientes = df_planilha.groupby(['Cliente']).value_counts()
df_metodo_pagamento_clientes = df_metodo_pagamento_clientes.reset_index()

grafico(x='Cliente', y='Desconto', hue='Método de Pagamento', data=df_metodo_pagamento_clientes,
        title='Métodos de pagamento mais utilizados pelos clientes e descontos aplicados', xlabel='Cliente', ylabel='Desconto')